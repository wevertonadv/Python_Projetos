# importantando a biblioteca openpyxl que é responsavel por ler e escrever operações no excel
import openpyxl
from openpyxl import Workbook
# importando a bilioteca pandas para let a planilha
import pandas as pd
tabela = pd.read_excel('disciplinas.xlsx')



for i, disciplina in enumerate(tabela['Disciplina']):
    # Armazene o objeto da classe Workbook em uma variável  
    wrkbk = openpyxl.Workbook()
    # Criando uma nova planilha
    sh = wrkbk.create_sheet("Nome do aluno", 0)
    # Definindo o valor na linha 2 e coluna 3
    sh.cell(row=1,column=1).value = 'Nome do aluno '
    # Para salvar a pasta de trabalho
    wrkbk.save(f'{disciplina}.xlsx')